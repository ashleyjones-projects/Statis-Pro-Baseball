# -*- coding: utf-8 -*-
"""
Created on Wed Jun 21 21:57:02 2017

@author: w9641432
"""

#%%
pitcher_dict={}
#Lag=['ARI','ATL','BAL','BOS','CHN','CHA','CIN','CLE','COL','DET','HOU','KCA','LAA','LAN','MIA','MIL','MIN','NYN','NYA','OAK','PHI','PIT','SLN','SDN','SFN','SEA','TBA','TEX','TOR','WAS']
ari_batters={};atl_batters={}; bal_batters={};bos_batters={};chn_batters={};\
cha_batters={};cin_batters={};cle_batters={};col_batters={};det_batters={};\
hou_batters={};kca_batters={};laa_batters={};lan_batters={};mia_batters={};\
mil_batters={};min_batters={};nyn_batters={};nya_batters={};oak_batters={};\
phi_batters={};pit_batters={};sln_batters={};sdn_batters={};sfn_batters={};\
sea_batters={};tba_batters={};tex_batters={};tor_batters={};was_batters={}

pitcher=[]
outpath='C:/Users/w9641432/Documents/Work/Python_learning/StatisProBaseball/teams/2016/'
import numpy as np
import openpyxl

# PULL OUT BIOS

wb = openpyxl.load_workbook('Master.xlsx')
sheet = wb.get_sheet_by_name('Master')

# PULL OUT FIELDING STATS

wb1 = openpyxl.load_workbook('Fielding.xlsx')
sheet1 = wb1.get_sheet_by_name('Fielding')

# PULL OUT FIELDING STATS FOR OUT FIELD PLAYERS

wb2 = openpyxl.load_workbook('FieldingOFsplit.xlsx')
sheet2 = wb2.get_sheet_by_name('FieldingOFsplit')

# PULL OUT PITCHING STATS

wb3 = openpyxl.load_workbook('pitching.xlsx')
sheet3 = wb3.get_sheet_by_name('Pitching')

#%%#%% Start extracting data
names=[]
all_names=[]
year=[]
stint=[]
team=[]
conf=[]
stats=[]
for rowOfCellObjects in sheet3['A44141':'A44964']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            all_names.append(cellObj.value)
for rowOfCellObjects in sheet3['A44141':'A44964']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            names.append(cellObj.value);            
            #print(cellObj.coordinate, cellObj.value);
for rowOfCellObjects in sheet3['B44141':'B44964']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            year.append(cellObj.value)
for rowOfCellObjects in sheet3['C44141':'C44964']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            stint.append(cellObj.value)
for rowOfCellObjects in sheet3['D44141':'D44964']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            team.append(cellObj.value)
for rowOfCellObjects in sheet3['E44141':'E44964']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            conf.append(cellObj.value)
for rowOfCellObjects in sheet3['F44141':'AD44964']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            stats.append(cellObj.value);
                        
#%% Load up hits chart
singlesP=[]

wb4 = openpyxl.load_workbook('hits_chart.xlsx')

sr = wb4.get_sheet_by_name('Pitcher_hits')


for rowOfCellObjects in sr['A1':'C22']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            singlesP.append(cellObj.value)

singlesP = np.reshape(singlesP,(22,3))
                              
#%%
# Find players over multiple years and collect stats


x = np.reshape(stats,(len(all_names),25))
 
for i in range(0,10): #len(names)
    data=[]
    n = names[i]
    obrpct=[]
    ts=[]
    avg=[]
    I=[]
    for j in range(0, len(all_names)):
        if n in all_names[j]:
            data.append(x[j])
            ts.append(team[j])
            #print(x[j][7]/x[j][24])
        else:dummy=1     
        
        avg=[float(sum(l))/len(l) for l in zip(*data)] # Finds the mean in a matrix
        era = np.float32(avg)
        avg=np.int32(avg)                            # Set to integers
        
        
    if len(ts)< 2:
        teams = ts
    else: 
        teams = ts[-1]

 #%%      
        # Extract bios
        
    import extract_bios
          
    bios = extract_bios.extract_bios(sheet,n)
    print('bios extracted for ', n)
        
        #%%
        # Extract Fielding attributes
        
    import extract_fielding
    
    Fielding = extract_fielding.extract_fielding(sheet1,sheet2,n,np)
    print('fielding extracted')   

#%%

    if (avg[7])>10: 
        
        # Work out PB from ERA, PB= 2 - ?
        
        if era[14]>5.0:
            pb=5
        elif era[14] >= 4.0 and era[14] < 5.0:
            pb=6
        elif era[14] >= 3.2 and era[14] < 4.0:
            pb=7
        elif era[14] >= 2.5 and era[14] < 3.2: 
            pb=8
        elif era[14] >= 1.9 and era[14] < 2.5:
            pb=9
        elif era[14] < 1.9:
            pb=10
        else:
            dummy=1
            
        if pb == 5 and avg[0]>12:
            pb=6
        elif pb == 6 and avg[0]>20:                  
            pb=7
        elif pb == 7 and avg[1]>20:                  
            pb=6
        elif pb == 8 and avg[1]>10:                  
            pb=7
        else:
            dummy=1
            
        # Work out CD (Cluthc defence)
        
        gidp = avg[7]/avg[24] # Outs/GIDP
        
        if gidp > 80:
            cd=0 
        elif gidp > 60 and gidp <=80:
            cd=1
        elif gidp > 40 and gidp <=60:
            cd=2
        elif gidp > 20 and gidp <=40:
            cd=3
        elif gidp <= 20:
            cd=4
        else:
            dummy=1
            
        # Start rating SR 
        
        if avg[3]>0:
            sr=np.int32((era[14]*2.0) + ((avg[8]+avg[11])/avg[3]))
        else:
            sr=0
            
        # Start rating RR 
        
        if (avg[2]-avg[3])>0:
            if sr > 0:
                rr=np.floor(sr/2)
            else:    
                rr=np.int32(((era[14]*2.0) + ((avg[8]+avg[11])/avg[2]))/2)
        else:
           rr=0 
        
        # Get number of singles for card

        import P_hits_per_inning

        singles = P_hits_per_inning.P_hits_per_inning(avg,pb)
        
        # Get number of walks and strikoouts for card
        import P_KBB_per_inning
        var = avg[12]/(avg[7]/3) # Ks per inning
        
        strikos = P_KBB_per_inning.P_KBB_per_inning(var,pb)         
                 
        var = avg[11]/(avg[7]/3) # walks per inning
        
        baseonb = P_KBB_per_inning.P_KBB_per_inning(var,pb)  
        
        # Calculate balks
        
        if avg[18] >= 6:
            
            balks = 2
        
        elif avg[18] > 0 and avg[18] < 6:
            
            balks = 1
            
        else:  
            balks = 0
            
        # Calculate Wild picthes
        
        if avg[16] >= 6:
            
            wildps = 2
        
        elif avg[16] > 0 and avg[16] < 6:
            
            wildps = 1
            
        else:  
            
            wildps = 0 
            
        # Calculate passed balls
        
        if baseonb >= 6:
            
            passbs = 2
        
        elif avg[18] > 3 and avg[18] < 6:
            
            passbs = 1
            
        else:  
            passbs = 0    
                          
#%%
        numbers = list(range(11,19)) + list(range(21,29)) + list(range(31,39)) + list(range(41,49)) + list(range(51,59))+ list(range(61,69)) + list(range(71,79)) + list(range(81,89))
# Card placement

  # SINGLES        
        
        # INFIELD SINGLES
        
        onebf = [11]
        infield = 1
        
        singles = singles - infield
        
        L = singlesP[singles-1][0]
        C = singlesP[singles-1][1]
        R = singlesP[singles-1][2]
           
        onebseven = [numbers[m] for m in range(infield,infield+L)]
        onebeight = [numbers[m] for m in range(infield+L,infield+L+C)]
        onebnine = [numbers[m] for m in range(infield+L+C,infield+L+C+R)]
        
        I = infield+L+C+R
        
        # ADD BALKS
        
        balkos = [numbers[m] for m in range(I,I+balks)] 
        
        I = I+balks
        
        # ADD STRIKES
        
        ks =  [numbers[m] for m in range(I,I+strikos)] 
        
        I = I+strikos
        
        # ADD WALKS
        if I + baseonb <= 64:
            ws = [numbers[m] for m in range(I,I+baseonb)]
            I = I+baseonb
        else:
            ws =  [numbers[i] for i in range(I,I+baseonb + (64-(I+baseonb)))]
            I = 64
        
        # ADD PASSSBALLS
        if I + passbs <= 64 and I <= 64:
            pbs = [numbers[m] for m in range(I,I+passbs)] 
            I = I+passbs
        else:    
            pbs =  [numbers[i] for i in range(I,I+passbs + (64-(I+passbs)))]
            I = 64
        
        # ADD WILD PICTHES
        if I + wildps <= 64 and I <= 64:
            wps =  [numbers[m] for m in range(I,I+wildps)] 
            I = I+wildps
        else:
            wps =  [numbers[i] for i in range(I,I+wildps + (64-(I+wildps)))]
            I = 64
        
        # ADD OUTS
        if I < 64:
            out = [numbers[i] for i in range(I,I+(len(numbers))-I)]
        else:
            out=[]
        print('Created card numbers')     
#%%    
       
 # Assign all indices in an array calling the actual action
        attributes = []
        for l in range(0,len(numbers)):
            if numbers[l] in onebf:
                attributes.append('1BF')
            elif numbers[l] in onebseven:
                attributes.append('1B7') 
            elif numbers[l] in onebeight:
                attributes.append('1B8')
            elif numbers[l] in onebnine:
                attributes.append('1B9') 
            elif numbers[l] in balkos:
                attributes.append('BKK') 
            elif numbers[l] in ks:
                attributes.append('KKK')
            elif numbers[l] in ws:
                attributes.append('WWW')
            elif numbers[l] in pbs:
                attributes.append('PBB')
            elif numbers[l] in wps:
                attributes.append('WPP')    
            elif numbers[l] in out:
                attributes.append('Out') 
        print(['created attributes, '+ str(len(attributes))])            
        
#%%        
        pitcher_list = Fielding[0],teams[0],bios[0],bios[1],bios[2],\
                         bios[4],bios[3],Fielding[1],Fielding[2],Fielding[3],\
                         Fielding[4],pb,cd,sr,rr,onebf,onebseven,onebeight,\
                         onebnine,balkos,ks,ws,pbs,wps,out,attributes,avg[3],avg[0],\
                         (avg[2]-avg[3]),avg[6],era[14]                                               
                          
        #print(Fielding[4])                
        import csv                  
        if 'ARI' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Arizona_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'ATL' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Atlanta_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'BOS' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Boston_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'BAL' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Baltimore_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'CHN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Chicago_N_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'CHA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Chicago_A_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'CIN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Cincinnati_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'CLE' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Cleveland_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'COL' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Colorado_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)         
        elif 'DET' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Detroit_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'HOU' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Houston_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'KCA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Kansas_city_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'LAA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Los_Angeles_A_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)         
        elif 'LAN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Los_Angeles_N_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'MIA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Miami_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'MIL' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Milwaukee_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'MIN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Minnasota_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list) 
        elif 'NYA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'New_Yotk_A_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'NYN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'New_York_N_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'OAK' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Oakland_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'PHI' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Philadelphia_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'PIT' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Pittsburgh_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'SLN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'St_Louis_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'SDN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'San_Diego_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'SFN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'San_Fransisco_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'SEA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Seattle_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'TBA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Tampa_Bay_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'TEX' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Texas_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)
        elif 'TOR' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Toronto_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)         
        elif 'WAS' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Washington_p.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(pitcher_list)        
        else:
            dummy=1