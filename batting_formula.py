# -*- coding: utf-8 -*-
"""
Created on Tue Aug 22 10:23:48 2017

@author: w9641432
"""

# -*- coding: utf-8 -*-
"""
Created on Tue Jun 20 14:27:48 2017

@author: w9641432
"""

#%%
class Batting:
    
    Version = "Ver 1.0 Rev 0"
    
    def __init__(player,code,team,Fname,Lname,nation,throw,Bat,posit,\
                 games,err,arm,obr,sp,cd,sac,hnr,onebf,onebs,onebe,onebn,\
                 twobs,twobe,twobn,threebe,hr,k,w,hbp,out,atts):
        
        player.Code = code
        player.Team = team
        player.Firstname = Fname
        player.Lastname = Lname
        player.Nationality = nation
        player.Throws = throw           
        player.Bats = Bat
        player.Positions = posit
        player.Games = games
        player.Errors = err           
        player.Arm = arm
        player.OBR = obr
        player.SP = sp
        player.CD = cd
        player.SAC = sac
        player.HNR = hnr
        player.ONEBF = onebf
        player.ONEBSEVEN = onebs
        player.ONEBEIGHT = onebe
        player.ONEBNINE = onebn
        player.TWOBSEVEN = twobs
        player.TWOBEIGHT = twobe
        player.TWOBNINE = twobn
        player.TRIPLEBCENTRE = threebe
        player.HR = hr
        player.K = k
        player.W =w
        player.HBP = hbp
        player.OUT = out
        player.Attributes =atts
        
#%%
batting_dict_p={}
lag=['ARI','ATL','BAL','BOS','CHN','CHA','CIN','CLE','COL','DET','HOU','KCA','LAA','LAN','MIA','MIL','MIN','NYN','NYA','OAK','PHI','PIT','SLN','SDN','SFN','SEA','TBA','TEX','TOR','WAS']
lagF=('Arizona (N)','Atlanta (N)','Baltimore (A)','Boston (A)','Chicago (N)','Chicago (A)','Cincinnati (N)','Cleveland (A)','Colorado (N)','Detroit (A)','Miami (N)','Houston (A)','Kansas City (A)','Los Angeles (A)','Los Angeles (N)','Milwaukee (N)','Minnesota (A)','New York (N)','New York (A)','Oakland (A)','Philadelphia (N)','Pittsburgh (N)','St. Louis Cardinals (N)','San Diego (N)','San Francisco (N)','Seattle (A)','Tampa Bay (A)','Texas (A)','Toronto (A)','Washington (N)')
ari_batters_p={};atl_batters_p={}; bal_batters_p={};bos_batters_p={};chn_batters_p={};\
cha_batters_p={};cin_batters_p={};cle_batters_p={};col_batters_p={};det_batters_p={};\
hou_batters_p={};kca_batters_p={};laa_batters_p={};lan_batters_p={};mia_batters_p={};\
mil_batters_p={};min_batters_p={};nyn_batters_p={};nya_batters_p={};oak_batters_p={};\
phi_batters_p={};pit_batters_p={};sln_batters_p={};sdn_batters_p={};sfn_batters_p={};\
sea_batters_p={};tba_batters_p={};tex_batters_p={};tor_batters_p={};was_batters_p={}

batter=[]
outpath='C:/Users/w9641432/Documents/Work/Python_learning/StatisProBaseball/teams/2016/'
import numpy as np
import openpyxl

# PULL OUT BIOS
wb = openpyxl.load_workbook('Master.xlsx')
sheet = wb.get_sheet_by_name('Master')

# PULL OUT FIELDING STATS

wb1 = openpyxl.load_workbook('Fielding.xlsx')
sheet1 = wb1.get_sheet_by_name('Fielding')

# PULL OUT FIELDING STATS FOR OUT FIELD PLAYERS

wb2 = openpyxl.load_workbook('FieldingOFsplit.xlsx')
sheet2 = wb2.get_sheet_by_name('FieldingOFsplit')

# PULL OUT BATTING STATS

wb3 = openpyxl.load_workbook('batting.xlsx')
sheet3 = wb3.get_sheet_by_name('batting')

#%% Load up hits chart
singlesR=[]
singlesL=[]
doublesR=[]
doublesL=[]
wb4 = openpyxl.load_workbook('hits_chart.xlsx')

sr = wb4.get_sheet_by_name('singles_R')
sl = wb4.get_sheet_by_name('singles_L')
dr = wb4.get_sheet_by_name('doubles_R')
dl = wb4.get_sheet_by_name('doubles_L')

for rowOfCellObjects in sr['A1':'C38']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            singlesR.append(cellObj.value)
for rowOfCellObjects in sl['A1':'C38']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            singlesL.append(cellObj.value)            
for rowOfCellObjects in dr['A1':'C35']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            doublesR.append(cellObj.value)
for rowOfCellObjects in dl['A1':'C35']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            doublesL.append(cellObj.value)  

singlesR = np.reshape(singlesR,(38,3))
singlesL = np.reshape(singlesL,(38,3)) 
doublesR = np.reshape(doublesR,(35,3)) 
doublesL = np.reshape(doublesL,(35,3))             



#%%#%% Start extracting data
names=[]
all_names=[]
year=[]
stint=[]
team=[]
conf=[]
stats=[]
for rowOfCellObjects in sheet3['A101335':'A102817']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            all_names.append(cellObj.value)
for rowOfCellObjects in sheet3['A101335':'A102817']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            names.append(cellObj.value);            
            #print(cellObj.coordinate, cellObj.value);
for rowOfCellObjects in sheet3['B101335':'B102817']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            year.append(cellObj.value)
for rowOfCellObjects in sheet3['C101335':'C102817']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            stint.append(cellObj.value)
for rowOfCellObjects in sheet3['D101335':'D102817']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            team.append(cellObj.value)
for rowOfCellObjects in sheet3['E101335':'E102817']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            conf.append(cellObj.value)
for rowOfCellObjects in sheet3['F101335':'V102817']: # Data from 2014-2016, cell number
        for cellObj in rowOfCellObjects:
            stats.append(cellObj.value);
#%%
        # Extract Fielding attributes
        
namesF=[]
infPosit=[]
infstats=[]
infteam=[]
for rowOfCellObjects in sheet1['A134864':'A136815']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            namesF.append(cellObj.value)
for rowOfCellObjects in sheet1['D134864':'D136815']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            infteam.append(cellObj.value)
for rowOfCellObjects in sheet1['F134864':'F136815']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            infPosit.append(cellObj.value)
for rowOfCellObjects in sheet1['G134864':'R136815']: # Data from 2014-2016,cell number 
        for cellObj in rowOfCellObjects:
            infstats.append(cellObj.value)          
            
#%%
# Find players over multiple years and collect stats


x = np.reshape(stats,(len(all_names),17))
stats = np.reshape(infstats,(len(namesF),12))
for i in range(0,100): #len(names)
    data=[]
    n = names[i]
    obrpct=[]
    ts=[]
    avg=[]
    I=[]
    for j in range(0, len(all_names)):
        if n in all_names[j]:
            data.append(x[j])
            ts.append(team[j])
            #print(x[j][7]/x[j][24])
        else:dummy=1     
        
        avg=[float(sum(l))/len(l) for l in zip(*data)] # Finds the mean in a matrix
    data=[]
    if len(ts)< 2:
        teams = ts
    else: 
        teams = ts[-1]
 #%%      
        # Extract bios
        
    import extract_bios
          
    bios = extract_bios.extract_bios(sheet,n)
    print('bios extracted for ', n)
        
        #%%
        # Extract Fielding attributes
        
    import extract_fielding
    
    Fielding = extract_fielding.extract_fielding(sheet1,sheet2,n,np)
    print('fielding extracted')   
    
#%%

    if any(avg) and ('C' in Fielding[1] or '1B' in Fielding[1] or '2B' in Fielding[1] or 'SS' in Fielding[1] or '3B' in Fielding[1] or 'OF' in Fielding[1]) and avg[1]>30: 
        #1) Stolen bases rating
        
        if   avg[8]>=30:
              sp=[1,'A']
        elif avg[8]>=20 and avg[8]<30:
              sp=[2,'B']
        elif avg[8]>=10 and avg[8]<20:
              sp=[3,'C']
        elif avg[8]>=1 and  avg[8]<10:
              sp=[4,'D']
        else:
              sp=[5,'E']
              
        #2) OBR rating
              
        obrpct = avg[2]/(avg[3]+avg[10]+avg[12]+avg[13])
        
        if obrpct>0.4:
            obr=[1,'A']
        elif obrpct>=0.35 and obrpct<0.4:
            obr=[2,'B']
        elif obrpct>=0.30 and obrpct<0.35:
            obr=[3,'C']
        elif obrpct>=0.25 and obrpct<0.30:
            obr=[4,'D']
        elif obrpct<0.25:
            obr=[5,'E']
        else:
            obr=[5,'E']   
        
        if sp[0]<obr[0]: # Check to make sure OBR cant be less than SP
            obr[0]=sp[0]
        else:
            dummy=1
            
            
         # Create CD values, looks a GIDP
        
        if avg[16]>15:
              cd=2
        elif avg[16]>=10 and avg[16]<15:
              cd=1
        elif avg[16]<10:
              cd=0
              
        # Create SAC values, looks at
        
        if avg[15]>=8:
            sac=[1,'AA']
        elif avg[15]>=5 and avg[15]<8:
            sac=[2,'BB']
        elif avg[15]>=2 and avg[15]<5:
            sac=[3,'CC']
        elif avg[15]<2:
            sac=[4,'DD']
            
        # Create batting numbers, using 128 as a factor
        
        tpa = avg[1] + avg[10] + avg[12] + avg[13]
        fact = tpa/128
        
        if tpa>0:
        
           singles = np.int32(round(((avg[3]-avg[4]-avg[5]-avg[6])/fact)))-12
           doubles = np.int32(round(avg[4]/fact))
           triples = np.int32(round(avg[5]/fact))
           dingers = np.int32(round(avg[6]/fact))
           strikos = np.int32(round(avg[11]/fact))-12
           baseonb = np.int32(round((avg[10]+avg[12])/fact))-8
           hitbypi = np.int32(round(avg[13]/fact))
           
           if singles<0:
               singles=0
           elif strikos<0:
               strikos=0
           elif baseonb<0:
               baseonb=0
           
        
        else:
            
           singles=0
           doubles=0
           triples=0
           dingers=0
           strikos=0
           baseonb=0
           hitbypi=0;
           
           
        
        # Determine type of batter, power or normal
        if avg[6]>=30 and 'R' in bios[4]:
           cht='RP'
        elif avg[6]>=30 and 'L' in bios[4]:
           cht='LP'
        elif avg[6]>=30 and 'B' in bios[4]:
           cht='SP'
        elif avg[6]<30 and 'R' in bios[4]:
           cht='RN'   
        elif avg[6]<30 and 'L' in bios[4]:
           cht='LN'
        elif avg[6]<30 and 'B' in bios[4]:
           cht='SN'   
        #Determine BD type
        
        
        #   BD:0    BD:1    BD:2  Result of play
        #   11-24   11-28   11-42  Double to right center. All runners score.
        #   25-26   31-34   43-44  Triple to left center.
        #      27   35-37   45-48  Home Run to deep center.
        #   28-88   38-88   51-88  No Action occurs. Return to normal play. 
        
#        if avg[6]>=30:
#            BD=2
#        elif avg[6]>=25 and avg[6]<30:
#            BD=1
#        else:
#            BD=0
        
        
        
        if doubles>0:
            
            if doubles>0 and doubles <=10 :
                
                twbd=(doubles*2)-1
            
            elif doubles > 10 and doubles <=20:
                
                twbd=(doubles*2)-2
        else:
            
            twbd=1
             
        if triples==0:
             
            if obrpct>=0.35:
                
                trbd = 1
            else:
                trbd = 0
                
        elif triples>0 and triples <=10 :
                trbd=(triples*2)-1
        elif triples > 10 and triples <=20:
                trbd=(triples*2)-2
        else:
             trbd=0 
             
        if dingers == 0 :
        
            hrsbd = 1
            
        elif dingers>0 and dingers <=10:
        
            hrsbd = (dingers*2)-1
         
        elif dingers>10 and dingers <=20:
        
            hrsbd = (dingers*2)-1            
                    
             
        # Determine hit and run number
        if strikos<1:
            hnr = 2
        elif strikos>=1 and strikos <=2:
            hnr = 1
        else:
            hnr = 0;
        print('created formula')  
#%% 
        numbers = list(range(11,19)) + list(range(21,29)) + list(range(31,39)) + list(range(41,49)) + list(range(51,59))+ list(range(61,69)) + list(range(71,79)) + list(range(81,89))
# Card placement

  # SINGLES
        if singles>0:
            if 'A' in obr: 
            
                onebf = [11,12]
                infield=1
            
            elif 'B' in obr: 
            
                onebf = [11]
                infield=0
            
            else:
                onebf = []
                infield=0;
            
                  
            singles = singles - infield
        
        
            if 'R' in bios[3]:
            
                L = singlesR[singles-1][0]
                C = singlesR[singles-1][1]
                R = singlesR[singles-1][2]
           
                onebseven = [numbers[m] for m in range(infield,infield+L)]
                onebeight = [numbers[m] for m in range(infield+L,infield+L+C)]
                onebnine = [numbers[m] for m in range(infield+L+C,infield+L+C+R)]
            
                I = infield+L+C+R
            
            elif 'L' in bios[3] or 'B' in bios[3]:
                
                L = singlesL[singles-1][0]
                C = singlesL[singles-1][1]
                R = singlesL[singles-1][2]
           
                onebseven = [numbers[m] for m in range(infield,infield+L)]
                onebeight = [numbers[m] for m in range(infield+L,infield+L+C)]
                onebnine = [numbers[m] for m in range(infield+L+C,infield+L+C+R)]
                
                I = infield+L+C+R
        
        else:        
            onebseven=[]
            onebeight=[]
            onebnine=[]
            
            I=0   
        
   # Doubles
   
        
        if doubles>0:
            
            if 'R' in bios[3]:
                
                L = doublesR[doubles-1][0]
                C = doublesR[doubles-1][1]
                R = doublesR[doubles-1][2]
            
                twobseven = [numbers[m] for m in range(I,I+L)]
                twobeight = [numbers[m] for m in range(I+L,I+L+C)]
                twobnine = [numbers[m] for m in range(I+L+C,I+L+C+R)]
            
                I = I+L+C+R
            
            elif 'L' in bios[3] or 'B' in bios[3]:
                
                L = doublesL[doubles-1][0]
                C = doublesL[doubles-1][1]
                R = doublesL[doubles-1][2]
            
                twobseven = [numbers[m] for m in range(I,I+L)]
                twobeight = [numbers[m] for m in range(I+L,I+L+C)]
                twobnine = [numbers[m] for m in range(I+L+C,I+L+C+R)]
            
                I = I+L+C+R
            
            
        else:
            twobseven=[]
            twobeight=[]
            twobnine=[]
        
        
 # Calculate triples

        triplecentre = [numbers[i] for i in range(I,I+triples)] 
        
        I = I+triples
        
 # Calculate homeruns
 
        homerun = [numbers[i] for i in range(I,I+dingers)]
        
        I = I+dingers
        
 # calculate strikeouts
        if I + strikos <= 64:
            
            ks = [numbers[i] for i in range(I,I+strikos)]
        
            I = I + strikos
        else:
            
            ks = [numbers[i] for i in range(I,I+strikos + (64-(I+strikos)))]
            
            I = 64
            
 # Calculate walks
        if I + baseonb <= 64 and I <= 64: 
            
            ws =  [numbers[i] for i in range(I,I+baseonb)]
        
            I = I + baseonb
            
        else: 
            
            ws =  [numbers[i] for i in range(I,I+baseonb + (64-(I+baseonb)))]
        
            I = 64
 # Calculate hit by Pitch
        if I + hitbypi <= 64 and I <= 64:
        
            hbp =  [numbers[i] for i in range(I,I+hitbypi)] 
       
            I = I + hitbypi
            
        else:
            
            hbp =  [numbers[i] for i in range(I,I+hitbypi + (64-(I+baseonb)))]
            
            I = 64
        
 # Out numbers
        if I < 64:
            out = [numbers[i] for i in range(I,I+(len(numbers))-I)]
        else:
            out=[]
            
#%% create BD ratings
        I=0
        bdtwo = [numbers[m] for m in range(0,twbd)]
        I=I+twbd
        bdtre = [numbers[m] for m in range(I,I+trbd)]
        I=I+trbd
        bdhrs = [numbers[m] for m in range(I,I+hrsbd)]
        I=I+hrsbd
        non = [numbers[i] for i in range(I,I+(len(numbers))-I)]
        print('Created card numbers')    


#%%    
       
 # Assign all indices in an array calling the actual action
        attributes = []
        for l in range(0,len(numbers)):
            if numbers[l] in onebf:
                attributes.append('1BF')
            elif numbers[l] in onebseven:
                attributes.append('1B7') 
            elif numbers[l] in onebeight:
                attributes.append('1B8')
            elif numbers[l] in onebnine:
                attributes.append('1B9') 
            elif numbers[l] in twobseven:
                attributes.append('2B7')
            elif numbers[l] in twobeight:
                attributes.append('2B8') 
            elif numbers[l] in twobnine:
                attributes.append('2B9')
            elif numbers[l] in triplecentre:
                attributes.append('3B8') 
            elif numbers[l] in homerun:
                attributes.append('HRR')
            elif numbers[l] in ks:
                attributes.append('KKK') 
            elif numbers[l] in ws:
                attributes.append('WWW')
            elif numbers[l] in hbp:
                attributes.append('HBP') 
            elif numbers[l] in out:
                attributes.append('Out') 
                
            # ADD bd Attributes
            attributes_bd=[]
        for l in range(0,len(numbers)):
            if numbers[l] in bdtwo:
                attributes_bd.append('2DD')
            elif numbers[l] in bdtre:
                attributes_bd.append('3DD') 
            elif numbers[l] in bdhrs:
                attributes_bd.append('HDD')
            elif numbers[l] in non:
                attributes_bd.append('Out') 
            
        print(['created attributes, '+ str(len(attributes))])      
                
#%%
# APPEND ALL ATTRIBUTES AND STATS TO THE CLASS BATTING 
        

        
        
#(player,code,team,Fname,Lname,nation,throw,Bat,posit,\
#                 games,err,arm,obr,sp,cd,sac,hnr,onebf,onebs,onebe,onebn,\
#                 twobs,twobe,twobn,threebe,hr,k,w,hbp,out,atts):          
        
        batter = Batting(Fielding[0],ts,bios[0],bios[1],bios[2],\
                         bios[4],bios[3],Fielding[1],Fielding[2],Fielding[3],\
                         Fielding[4],obr,sp,cd,sac,hnr,onebf,onebseven,onebeight,\
                         onebnine,twobseven,twobeight,twobnine,triplecentre,homerun,\
                         ks,ws,hbp,out,attributes) 
        print('created batter profile')
        
#%%

# MAKE PLAYER DICTIONARY AND FRO TEAMS

        
        
        batting_dict_p[n]=batter
        print('stored profile in dictionary')


#        if 'ARI' in batter.Team:
#            ari_batters_p[n]=batter
#        elif 'ATL' in batter.Team:
#            atl_batters_p[n]=batter
#        elif 'BAL' in batter.Team:
#            bal_batters_p[n]=batter
#        elif 'BOS' in batter.Team:
#            bos_batters_p[n]=batter
#        elif 'CHN' in batter.Team:
#            chn_batters_p[n]=batter
#        elif 'CHA' in batter.Team:
#            cha_batters_p[n]=batter
#        elif 'CIN' in batter.Team:
#            cin_batters_p[n]=batter
#        elif 'CLE' in batter.Team:
#            cle_batters_p[n]=batter
#        elif 'COL' in batter.Team:
#            col_batters_p[n]=batter
#        elif 'DET' in batter.Team:
#            det_batters_p[n]=batter
#        elif 'HOU' in batter.Team:
#            hou_batters_p[n]=batter
#        elif 'KCA' in batter.Team:
#            kca_batters_p[n]=batter
#        elif 'LAA' in batter.Team:
#            laa_batters_p[n]=batter
#        elif 'LAN' in batter.Team:
#            lan_batters_p[n]=batter
#        elif 'MIA' in batter.Team:
#            mia_batters_p[n]=batter
#        elif 'MIL' in batter.Team:
#            mil_batters_p[n]=batter
#        elif 'MIN' in batter.Team:
#            min_batters_p[n]=batter
#        elif 'NYA' in batter.Team:
#            nya_batters_p[n]=batter
#        elif 'NYN' in batter.Team:
#            nyn_batters_p[n]=batter
#        elif 'OAK' in batter.Team:
#            oak_batters_p[n]=batter
#        elif 'PHI' in batter.Team:
#            phi_batters_p[n]=batter
#        elif 'PIT' in batter.Team:
#            pit_batters_p[n]=batter
#        elif 'SLN' in batter.Team:
#            sln_batters_p[n]=batter
#        elif 'SDN' in batter.Team:
#            sdn_batters_p[n]=batter
#        elif 'SFN' in batter.Team:
#            sfn_batters_p[n]=batter
#        elif 'SEA' in batter.Team:
#            sea_batters_p[n]=batter
#        elif 'TBA' in batter.Team:
#            tba_batters_p[n]=batter
#        elif 'TEX' in batter.Team:
#            tex_batters_p[n]=batter
#        elif 'TOR' in batter.Team:
#            tor_batters_p[n]=batter
#        elif 'WAS' in batter.Team:
#            was_batters_p[n]=batter
#        else:dummy=1             
#                              
#%%
        
        batter_list = Fielding[0],ts[0],bios[0],bios[1],bios[2],\
                         bios[4],bios[3],Fielding[1],Fielding[2],Fielding[3],\
                         Fielding[4],obr,sp,cd,sac,hnr,onebf,onebseven,onebeight,\
                         onebnine,twobseven,twobeight,twobnine,triplecentre,homerun,\
                         ks,ws,hbp,out,attributes,avg[0],avg[6],round(np.float(avg[3]/avg[1]),3),\
                         bdtwo,bdtre,bdhrs,attributes_bd,cht                            
        print(Fielding[4])                
        import csv                  
        if 'ARI' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Arizona.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'ATL' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Atlanta.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'BOS' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Boston.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'BAL' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Baltimore.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'CHN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Chicago_N.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'CHA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Chicago_A.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'CIN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Cincinnati.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'CLE' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Cleveland.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'COL' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Colorado.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)         
        elif 'DET' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Detroit.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'HOU' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Houston.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'KCA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Kansas_city.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'LAA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Los_Angeles_A.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)         
        elif 'LAN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Los_Angeles_N.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'MIA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Miami.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'MIL' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Milwaukee.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'MIN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Minnasota.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list) 
        elif 'NYA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'New_Yotk_A.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'NYN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'New_York_N.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'OAK' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Oakland.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'PHI' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Philadelphia.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'PIT' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Pittsburgh.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'SLN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'St_Louis.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'SDN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'San_Diego.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'SFN' in teams[0]:
            print(teams[0]) 
            with open(outpath+'San_Fransisco.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'SEA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Seattle.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'TBA' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Tampa_Bay.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'TEX' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Texas.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)
        elif 'TOR' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Toronto.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)         
        elif 'WAS' in teams[0]:
            print(teams[0]) 
            with open(outpath+'Washington.csv', 'a', newline='') as csv_file:
                wr = csv.writer(csv_file, delimiter=',')
                wr.writerow(batter_list)        
        else:
            dummy=1


                        
        
               